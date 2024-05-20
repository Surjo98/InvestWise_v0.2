import numpy as np
import pandas as pd
import os
import re
from openpyxl import load_workbook, Workbook

from sklearn.impute import KNNImputer

# Path to the Trendlyne folder
folder_path_tl = "data\Trendlyne"
folder_path_tt = "data\Tickertape"

# Split all_stocks.csv into sector-wise .csv files
file_path = os.path.join(folder_path_tt, 'all_stocks.csv')
df = pd.read_csv(file_path)

# Dictionary to map subsectors to sectors
subsector_to_sector = {
    'Consumer Discretionary': ['Auto Parts', 'Tires & Rubber', 'Four Wheelers', 'Three Wheelers', 'Two Wheelers', 'Cycles', 'Education Services', 'Wellness Services', 'Hotels, Resorts & Cruise Lines', 'Restaurants & Cafes', 'Theme Parks & Gaming', 'Tour & Travel Services', 'Home Electronics & Appliances', 'Home Furnishing', 'Housewares', 'Retail - Apparel', 'Retail - Department Stores', 'Retail - Online', 'Retail - Speciality', 'Apparel & Accessories', 'Footwear', 'Precious Metals, Jewellery & Watches', 'Textiles', 'Animation'],
    'Communication Services': ['Advertising', 'Cable & D2H', 'Movies & TV Serials', 'Publishing', 'Radio', 'Theatres', 'TV Channels & Broadcasters', 'Online Services', 'Telecom Equipments', 'Telecom Infrastructure', 'Telecom Services'],
    'Consumer Staples': ['Alcoholic Beverages', 'Soft Drinks', 'Tea & Coffee', 'Agro Products', 'FMCG - Foods', 'Packaged Foods & Meats', 'Seeds', 'Sugar', 'FMCG - Household Products', 'FMCG - Personal Products', 'FMCG - Tobacco'],
    'Energy': ['Oil & Gas - Equipment & Services', 'Oil & Gas - Exploration & Production', 'Oil & Gas - Refining & Marketing', 'Oil & Gas - Storage & Transportation'],
    'Financials': ['Private Banks', 'Public Banks', 'Asset Management', 'Investment Banking & Brokerage', 'Stock Exchanges & Ratings', 'Consumer Finance', 'Diversified Financials', 'Specialized Finance', 'Insurance', 'Home Financing', 'Payment Infrastructure'],
    'Health Care' : ['Biotechnology', 'Health Care Equipment & Supplies', 'Hospitals & Diagnostic Centres', 'Labs & Life Sciences Services', 'Pharmaceuticals'],
    'Industrials' : ['Airlines', 'Building Products - Ceramics', 'Building Products - Glass', 'Building Products - Granite', 'Building Products - Laminates', 'Building Products - Pipes', 'Business Support Services', 'Stationery', 'Construction & Engineering', 'Batteries', 'Cables', 'Electrical Components & Equipments', 'Heavy Electrical Equipments', 'Conglomerates', 'Agricultural & Farm Machinery', 'Heavy Machinery', 'Industrial Machinery', 'Rail', 'Shipbuilding', 'Tractors', 'Trucks & Buses', 'Employment Services', 'Airports', 'Dredging', 'Logistics', 'Ports', 'Roads', 'Renewable Energy Equipment & Services', 'Building Products - Prefab Structures', 'Commodities Trading', 'Aerospace & Defense Equipments'],
    'Information Technology': ['Communication & Networking', 'Electronic Equipments', 'Technology Hardware', 'IT Services & Consulting', 'Outsourced services', 'Software Services'],
    'Materials': ['Commodity Chemicals', 'Diversified Chemicals', 'Fertilizers & Agro Chemicals', 'Paints', 'Plastic Products', 'Specialty Chemicals', 'Cement', 'Packaging', 'Iron & Steel', 'Metals - Aluminium', 'Metals - Coke', 'Metals - Copper', 'Metals - Diversified', 'Metals - Lead', 'Mining - Coal', 'Mining - Copper', 'Mining - Diversified', 'Mining - Iron Ore', 'Mining - Manganese', 'Paper Products', 'Wood Products'],
    'Real Estate': ['Real Estate'],
    'Utilities' : ['Power Infrastructure', 'Power Transmission & Distribution', 'Gas Distribution', 'Power Trading & Consultancy', 'Power Generation', 'Renewable Energy', 'Water Management'],
    'ETF': ['Gold', 'Equity', 'Debt']
    # Add more sectors and their corresponding subsectors as needed
}

# Reverse the dictionary to map each subsector to its sector
subsector_to_sector_reverse = {subsector: sector for sector, subsectors in subsector_to_sector.items() for subsector in subsectors}

# Map the subsector to sector
df['Sector'] = df['Sub-Sector'].map(subsector_to_sector_reverse)

# Split the DataFrame into multiple DataFrames based on the 'Sector' column
sector_dfs = {sector: group.drop(columns=['Sector']) for sector, group in df.groupby('Sector')}

# Save the DataFrames to CSV
for sector, sector_df in sector_dfs.items():
    if not sector_df.empty:
        file_path = os.path.join(folder_path_tt, f'{sector}.csv')
        sector_df.to_csv(file_path, index=False)

# List to store dataframes
dfs = []
for file in os.listdir(folder_path_tl):
    if file.endswith(".xlsx"):

        file_path = os.path.join(folder_path_tl, file)
        df = pd.read_excel(file_path)
        df.drop(columns=['Stock Name', 'BSE code', 'ISIN', 'Current Price', 'Industry Name'], inplace=True)
        df.rename(columns={'NSE code': 'Ticker'}, inplace=True)
        dfs.append(df)

df_tl = pd.concat(dfs, ignore_index=True)


dfs = []
for file_name in os.listdir(folder_path_tt):
    if file_name.endswith('.csv'):

        file_path = os.path.join(folder_path_tt, file_name)
        df = pd.read_csv(file_path)
        df.insert(2, 'Sector', file_name.split('.')[0])
        dfs.append(df)

df_tt = pd.concat(dfs, ignore_index=True)

# Concatenate sub-dfs into final df
df = pd.merge(df_tt, df_tl, on='Ticker', how='inner')

# Create a new Excel writer object
with pd.ExcelWriter('Sector_Universe_Data.xlsx') as writer:
    # Iterate over unique sectors
    for sector in df['Sector'].unique():
        sector_df = df[df['Sector'] == sector]
        sector_df.to_excel(writer, sheet_name=sector, index=False)


# data files
input_file = 'Sector_Universe_Data.xlsx'
output_file = 'Sector Analysis New.xlsx'

# create a new workbook to store data
wb_out = Workbook()
wb_in = load_workbook(input_file)

for sheet_name in wb_in.sheetnames:

    df = pd.read_excel(input_file, sheet_name=sheet_name)

    def categorize_market_cap(row):
        if row['Market Cap'] > 82000:
            return 'Large'
        elif 26000 < row['Market Cap'] <= 82000:
            return 'Medium'
        else:
            return 'Small'

    df['Market Cap'] = df.apply(categorize_market_cap, axis=1)

    df.columns = df.columns.str.strip()

    def clean_column_name(column_name):
        # Keep only letters, numbers, spaces, and '/'
        cleaned_name = re.sub(r'[^a-zA-Z0-9 / -]', '', column_name)
        return cleaned_name

    # Applying the cleaning function to column names
    df.columns = [clean_column_name(col) for col in df.columns]

    columns_to_fill_zero = [
    'Insider Trades - 3M Cumulative',
    'Bulk Deals - 3M Cumulative',
    'FII Holding Change3M',
    'DII Holding Change3M',
    'MF Holding Change3M',
    'Promoter Holding Change3M',
    ]
    df[columns_to_fill_zero].fillna(0, inplace=True)

    for col in df.columns:
        if df[col].isna().sum() > 0.25 * len(df):
            df[col].fillna(0, inplace=True)


    small_df = df[df['Market Cap'] == 'Small']
    medium_df = df[df['Market Cap'] == 'Medium']
    large_df = df[df['Market Cap'] == 'Large']

    def apply_knn_imputation(segment_df):
        # excluding non-numeric columns
        numeric_cols = segment_df.select_dtypes(include=[np.number]).columns
        
        try:
            imputer = KNNImputer(n_neighbors=5)
            segment_df[numeric_cols] = imputer.fit_transform(segment_df[numeric_cols])
        except Exception as e:
            print(f"An error occurred during KNN imputation: {e}")
            segment_df[numeric_cols] = segment_df[numeric_cols].fillna(0)
        
        return segment_df

    # Apply KNN Imputation to each segment
    small_df_imputed = apply_knn_imputation(small_df)
    medium_df_imputed = apply_knn_imputation(medium_df)
    large_df_imputed = apply_knn_imputation(large_df)

    df = pd.concat([small_df_imputed, medium_df_imputed, large_df_imputed])

    def create_inverse(df, column):
        
        # Calculate inverse of specified column, handling zeros
        inverse_column = 1 / df[column].replace(0, np.nan)
        max_value = inverse_column.max()
        inverse_column = inverse_column.apply(lambda x: abs(x) + max_value if x < 0 else x)
        inverse_column.fillna(max_value + 1, inplace=True)
        
        df[column] = inverse_column
        return df


    # ## Feature Engineering

    df['Altman Zscore'] = df['Altman Zscore'] - 3
    df['Fundamentals'] = df['Piotroski Score'] * df['Altman Zscore']
    df.drop(columns=['Piotroski Score', 'Altman Zscore'], inplace=True)


    investor_cols = ['FII Holding Change3M', 'DII Holding Change3M', 'MF Holding Change3M', 'Promoter Holding Change3M', 'Insider Trades - 3M Cumulative', 'Bulk Deals - 3M Cumulative']
    df['Investor Sentiment'] = df[investor_cols].mean(axis=1)
    insider_cols = ['Insider Trades - 3M Cumulative', 'Promoter Holding Change3M']
    df['Promoter Sentiment'] = df[insider_cols].mean(axis=1)
    df.drop(columns=investor_cols+insider_cols, inplace=True)


    price_cols = ['Price / Sales', 'Price / CFO']
    df['Pricing'] = df[price_cols].mean(axis=1)*df['PE Ratio']*df['PB Ratio']
    df.drop(columns=price_cols+['PE Ratio', 'PB Ratio'], inplace=True)


    valuation_cols = ['EV / Invested Capital','EV / Revenue Ratio', 'EV / Free Cash Flow', 'EV/EBITDA Ratio']
    df['Valuation'] = df[valuation_cols].mean(axis=1)
    df.drop(columns=valuation_cols, inplace=True)


    profitability_cols = ['Return on Investment', 'ROCE', 'Return on Equity', 'Return on Assets', 'Dividend Yield']
    df['Profitability'] = df[profitability_cols].mean(axis=1)*df['Net Profit Margin']
    df.drop(columns=profitability_cols+['Net Profit Margin'], inplace=True)


    df = create_inverse(df, 'Cash Conversion Cycle')
    turnover_cols = ['Interest Coverage Ratio', 'Asset Turnover Ratio', 'Inventory Turnover Ratio', 'Working Capital Turnover Ratio']
    df['Business Turnover'] = df[turnover_cols].mean(axis=1)*df['Current Ratio']*df['Cash Conversion Cycle']
    df.drop(columns=turnover_cols +['Current Ratio', 'Cash Conversion Cycle'], inplace=True)


    future_eps = df['1Y Forward EPS Growth'] - df['1Y Historical EPS Growth']
    future_rev = df['1Y Forward Revenue Growth'] - df['1Y Historical Revenue Growth']
    df['Future Growth'] = future_eps + future_rev
    df.drop(columns=['1Y Forward EPS Growth', '1Y Historical EPS Growth', '1Y Forward Revenue Growth', '1Y Historical Revenue Growth'], inplace=True)


    # Define weights for each rank
    weights = {
        'Rank1': 0.2,
        'Rank2': 0.1,
        'Rank3': 0.15,
        'Rank4': 0.1,
        'Rank5': 0.1,
        'Rank6': 0.05,
        'Rank7': 0.1,
        'Rank8': 0.2
    }

    # Group by 'Market Cap' categories
    grouped = df.groupby('Market Cap')

    # Initialize an empty list to store the results
    result_dfs = []

    # Iterate over each group
    for name, group in grouped:
        # Calculate ranks within each group
        group['Rank1'] = group['Fundamentals'].rank(method='min', ascending=False)
        group['Rank2'] = group['Investor Sentiment'].rank(method='min', ascending=False)
        group['Rank3'] = group['Promoter Sentiment'].rank(method='min', ascending=False)
        group['Rank4'] = group['Pricing'].rank(method='min', ascending=True)
        group['Rank5'] = group['Valuation'].rank(method='min', ascending=True)
        group['Rank6'] = group['Profitability'].rank(method='min', ascending=False)
        group['Rank7'] = group['Business Turnover'].rank(method='min', ascending=False)
        group['Rank8'] = group['Future Growth'].rank(method='min', ascending=False)
        
        # Calculate weighted rank
        group['Weighted_rank'] = group[list(weights.keys())].mul(list(weights.values())).sum(axis=1)
        
        # Calculate final rank
        group['Rank'] = group['Weighted_rank'].rank(method='min', ascending=True)
        
        # Reorder columns
        columns_order = ['Rank', 'Name', 'Ticker', 'Sector', 'Sub-Sector', 'Market Cap', 'Fundamentals', 'Rank1', 'Investor Sentiment', 'Rank2', 'Promoter Sentiment', 'Rank3', 'Pricing', 'Rank4', 'Valuation', 'Rank5', 'Profitability', 'Rank6', 'Business Turnover', 'Rank7', 'Future Growth', 'Rank8', 'Weighted_rank']
        group = group[columns_order]
        
        # Append the result to the list
        result_dfs.append(group)

    # Concatenate the results
    df = pd.concat(result_dfs)
    
    # ## Final Model
    df.sort_values(by='Rank', ascending=True, inplace=True)

    # Create a new sheet called 'Healthcare'
    ws = wb_out.create_sheet(title=sheet_name)

    # Group the data by Market Cap and save them in different dataframes
    grouped = df.groupby('Market Cap')

    # Initialize row counter
    row_counter = 4

    # Loop through each group
    for name, group in grouped:
        # Write the headers starting from the fourth row
        ws.cell(row=row_counter + 1, column=1)
        ws.append(group.columns.tolist())  # Write headers
        
        # Write the data
        for row in group.itertuples(index=False):
            ws.append(list(row))
        
        # Add space of 4 rows between each dataframe
        row_counter = ws.max_row + 1

    # Save the workbook
    wb_out.save(output_file)

    print(f"Data has been saved for {sheet_name} in the Excel file.")