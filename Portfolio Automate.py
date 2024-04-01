import numpy as np
import pandas as pd
import openpyxl
import yfinance as yf
from scipy.optimize import minimize

# Read the DataFrame containing client names
name_df = pd.read_excel('PortfolioManagement.xlsx', sheet_name='Clients')

# Extract client names
client_names = name_df['Name']

# Load the workbook
wb = openpyxl.load_workbook('PortfolioManagement.xlsx')

# Iterate over each client name
for client_name in client_names:
    # Access the sheet corresponding to the client name
    if client_name in wb.sheetnames:
        ws = wb[client_name]
        
        # Assuming the first dataframe is from A1 to B16
        stats_info = {}
        for row in ws.iter_rows(min_row=1, max_row=16, min_col=1, max_col=2, values_only=True):
            stats_info[row[0]] = row[1]
        
        # Specify the target date and stock
        benchmark_stock = stats_info.get("Benchmark", None)
        start_date = stats_info.get("Start Date", None)
        current_date = stats_info.get("Current Date", None)
        
        # Check if benchmark stock and start date are provided
        if benchmark_stock is not None and start_date is not None:
            try:
                # Download benchmark data from Yahoo Finance
                benchmark_data = yf.download(benchmark_stock, start=start_date, end=current_date)
                
                # Calculate benchmark return
                benchmark_return = (benchmark_data["Adj Close"][-1] - benchmark_data["Adj Close"][0]) / benchmark_data["Adj Close"][0]
                
                # Update stats_info with benchmark return
                stats_info["Benchmark Return"] = benchmark_return
                
                if "Benchmark Return" in [cell.value for cell in ws["A"]]:
                    for cell in ws["A"]:
                        if cell.value == "Benchmark Return":
                            ws[f"B{cell.row}"] = benchmark_return
                            break
                
                print(f"Successfully updated benchmark for {client_name}.")
                # wb.save("PortfolioManagement.xlsx")
            except Exception as e:
                print(f"Error occurred while processing benchmark for {client_name}: {e}")
        else:
            print(f"Missing values in sheet {client_name}.")

        # Load the DataFrame from the Excel sheet
        df = pd.read_excel('PortfolioManagement.xlsx', sheet_name=client_name, skiprows=18)
        df = df[df['Weights'] > 0]
                  

        # Define the objective function to maximize Treynor ratio
        def treynor_max(weights, returns, beta, risk_free_rate):
            portfolio_return = np.dot(returns, weights)
            portfolio_beta = np.dot(beta, weights)
            treynor_ratio = (portfolio_return - risk_free_rate) / portfolio_beta
            return -treynor_ratio  # Minimize the negative Treynor ratio

        # Extract necessary data
        returns = df['Actual Return']/df['Invested Amount']
        risk_free_rate = stats_info.get("Risk-Free Rate", None)
        beta = df["Beta"].values
        weights = df['Weights'].to_list()

        if not any(obj is None for obj in [returns, risk_free_rate, beta]):
            # Define constraints
            constraints = ({'type': 'eq', 'fun': lambda weights: np.sum(weights) - 1})

            # Define bounds for each variable (weight)
            bounds = [(0, 3/len(returns)) for _ in range(len(returns))]

            # Perform optimization
            result = minimize(treynor_max, weights, args=(returns, beta, risk_free_rate),
                            method='SLSQP', bounds=bounds, constraints=constraints)

            # Extract optimal weights
            optimal_weights = result.x

            # Assign optimal weights to the DataFrame
            df['Optimized Weights'] = optimal_weights

            print(f"Optimization successful for {client_name}.")
            
            # Write the optimized weights back to the Excel sheet
            for i, value in enumerate(optimal_weights):
                ws.cell(row=20 + i, column=13, value=value)

        else:
            print(f"Missing data for optimization in sheet {client_name}.")

    else:
        print(f"No sheet found for client {client_name}.")

# Save the workbook
wb.save("PortfolioManagement.xlsx")